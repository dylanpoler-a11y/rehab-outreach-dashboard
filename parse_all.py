"""
Parses all 4 data sources and produces a unified data.json for the dashboard:
1. View All (3).csv — primary company data with HQ addresses for geocoding
2. Working Sheet (8).csv — company-level data (addresses, ownership, states, etc.)
3. Grid view copy (1).csv — message-level outreach data (the complete picture)
4. Bradford_Pipeline_Dashboard (3).xlsx — post-intro pipeline deals
"""

import csv
import json
import re
import hashlib
import openpyxl
import os
import io
import time
import urllib.request
import urllib.parse
from collections import defaultdict

# ============================================================
# GEOCODING
# ============================================================
STATE_COORDS = {
    'AL': [32.806671, -86.791130], 'AK': [61.370716, -152.404419],
    'AZ': [33.729759, -111.431221], 'AR': [34.969704, -92.373123],
    'CA': [36.116203, -119.681564], 'CO': [39.059811, -105.311104],
    'CT': [41.597782, -72.755371], 'DE': [39.318523, -75.507141],
    'FL': [27.766279, -81.686783], 'GA': [33.040619, -83.643074],
    'HI': [21.094318, -157.498337], 'ID': [44.240459, -114.478828],
    'IL': [40.349457, -88.986137], 'IN': [39.849426, -86.258278],
    'IA': [42.011539, -93.210526], 'KS': [38.526600, -96.726486],
    'KY': [37.668140, -84.670067], 'LA': [31.169546, -91.867805],
    'ME': [44.693947, -69.381927], 'MD': [39.063946, -76.802101],
    'MA': [42.230171, -71.530106], 'MI': [43.326618, -84.536095],
    'MN': [45.694454, -93.900192], 'MS': [32.741646, -89.678696],
    'MO': [38.456085, -92.288368], 'MT': [46.921925, -110.454353],
    'NE': [41.125370, -98.268082], 'NV': [38.313515, -117.055374],
    'NH': [43.452492, -71.563896], 'NJ': [40.298904, -74.521011],
    'NM': [34.840515, -106.248482], 'NY': [42.165726, -74.948051],
    'NC': [35.630066, -79.806419], 'ND': [47.528912, -99.784012],
    'OH': [40.388783, -82.764915], 'OK': [35.565342, -96.928917],
    'OR': [44.572021, -122.070938], 'PA': [40.590752, -77.209755],
    'RI': [41.680893, -71.511780], 'SC': [33.856892, -80.945007],
    'SD': [44.299782, -99.438828], 'TN': [35.747845, -86.692345],
    'TX': [31.054487, -97.563461], 'UT': [40.150032, -111.862434],
    'VT': [44.045876, -72.710686], 'VA': [37.769337, -78.169968],
    'WA': [47.400902, -121.490494], 'WV': [38.491226, -80.954456],
    'WI': [44.268543, -89.616508], 'WY': [42.755966, -107.302490],
    'DC': [38.897438, -77.026817]
}

# Full state name to abbreviation
STATE_ABBREV = {v: k for k, v in {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'Washington DC'
}.items()}

FL_ZIP_REGIONS = {
    '320': [30.4, -84.3], '321': [29.2, -81.0], '322': [30.3, -81.7],
    '323': [30.3, -81.7], '324': [30.4, -86.6], '325': [30.4, -87.2],
    '326': [29.2, -82.1], '327': [28.5, -81.4], '328': [28.5, -81.4],
    '329': [28.1, -80.6], '330': [25.8, -80.2], '331': [25.8, -80.2],
    '332': [25.8, -80.2], '333': [26.1, -80.1], '334': [26.7, -80.1],
    '335': [27.8, -82.6], '336': [27.8, -82.6], '337': [27.3, -82.5],
    '338': [28.0, -82.0], '339': [26.6, -81.9], '340': [26.6, -81.9],
    '341': [26.1, -80.4], '342': [28.0, -82.5], '344': [29.0, -82.5],
    '346': [27.5, -82.5], '347': [28.2, -82.2], '349': [26.4, -80.1],
}


def get_coords(address, state_abbr, name):
    lat, lng = None, None
    zip_match = re.search(r'\b(\d{5})\b', address) if address else None

    if zip_match:
        prefix = zip_match.group(1)[:3]
        if state_abbr == 'FL' and prefix in FL_ZIP_REGIONS:
            lat, lng = FL_ZIP_REGIONS[prefix]
        elif state_abbr in STATE_COORDS:
            lat, lng = STATE_COORDS[state_abbr]
    elif state_abbr in STATE_COORDS:
        lat, lng = STATE_COORDS[state_abbr]

    if lat is not None:
        h = int(hashlib.md5(name.encode()).hexdigest()[:8], 16)
        lat += ((h % 1000) / 1000 - 0.5) * 0.8
        lng += (((h >> 10) % 1000) / 1000 - 0.5) * 0.8

    return lat, lng


GEOCODE_CACHE_FILE = '/Users/dylanpoler/Downloads/rehab_dashboard/geocode_cache.json'


def load_geocode_cache():
    if os.path.exists(GEOCODE_CACHE_FILE):
        with open(GEOCODE_CACHE_FILE, 'r') as f:
            return json.load(f)
    return {}


def save_geocode_cache(cache):
    with open(GEOCODE_CACHE_FILE, 'w') as f:
        json.dump(cache, f, indent=2)


def batch_geocode_census(addresses):
    """
    Batch geocode addresses using US Census Bureau Geocoding API.
    addresses: list of (id, address_str) tuples
    Returns: dict of id -> (lat, lng) or id -> None
    """
    results = {}
    # Census batch API accepts up to 10,000 records
    # Format: id, street, city, state, zip
    batch_size = 1000
    for batch_start in range(0, len(addresses), batch_size):
        batch = addresses[batch_start:batch_start + batch_size]
        print(f"  Geocoding batch {batch_start//batch_size + 1} ({len(batch)} addresses)...")

        # Build CSV for batch geocoding
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        for uid, addr in batch:
            # Parse address into components
            parts = [p.strip() for p in addr.split(',')]
            if len(parts) >= 4:
                # "Street, City, State ZIP, Country" or "Street, City, State, ZIP, Country"
                street = parts[0]
                city = parts[1] if len(parts) > 1 else ''
                # State and zip might be combined like "Georgia 30092"
                state_zip = parts[2] if len(parts) > 2 else ''
                sz_match = re.match(r'([A-Za-z\s]+?)\s*(\d{5})', state_zip)
                if sz_match:
                    state = sz_match.group(1).strip()
                    zipcode = sz_match.group(2)
                else:
                    state = state_zip
                    zipcode = ''
                writer.writerow([uid, street, city, state, zipcode])
            elif len(parts) >= 2:
                writer.writerow([uid, parts[0], parts[1] if len(parts) > 1 else '', '', ''])
            else:
                writer.writerow([uid, addr, '', '', ''])

        csv_data = csv_content.getvalue().encode('utf-8')

        try:
            url = 'https://geocoding.geo.census.gov/geocoder/locations/addressbatch'
            boundary = '----BatchBoundary'
            body = (
                f'--{boundary}\r\n'
                f'Content-Disposition: form-data; name="addressFile"; filename="addresses.csv"\r\n'
                f'Content-Type: text/csv\r\n\r\n'
            ).encode('utf-8') + csv_data + (
                f'\r\n--{boundary}\r\n'
                f'Content-Disposition: form-data; name="benchmark"\r\n\r\n'
                f'Public_AR_Current\r\n'
                f'--{boundary}--\r\n'
            ).encode('utf-8')

            req = urllib.request.Request(
                url,
                data=body,
                headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
                method='POST'
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                response_text = resp.read().decode('utf-8')

            # Parse response CSV
            for line in response_text.strip().split('\n'):
                if not line.strip():
                    continue
                # Response format: "id","input address","match","exact/non-exact","matched address","coords","tiger id","side"
                row_reader = csv.reader(io.StringIO(line))
                for row in row_reader:
                    if len(row) >= 6 and row[2].strip().lower() == 'match':
                        uid = row[0].strip().strip('"')
                        coords_str = row[5].strip().strip('"')
                        if coords_str:
                            try:
                                lng_str, lat_str = coords_str.split(',')
                                results[uid] = (float(lat_str.strip()), float(lng_str.strip()))
                            except (ValueError, IndexError):
                                pass

            print(f"    Got {sum(1 for b in batch if b[0] in results)} matches")
        except Exception as e:
            print(f"    Geocoding error: {e}")

        # Rate limit
        if batch_start + batch_size < len(addresses):
            time.sleep(1)

    return results


# ============================================================
# PARSE VIEW ALL CSV (primary address source)
# ============================================================
def parse_view_all(filepath):
    """Returns dict keyed by normalized company name with address and allStates."""
    companies = {}
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r.get('Name', '').strip()
            if not name:
                continue
            key = name.lower().strip()
            address = r.get('HQ Address', '').strip()
            all_states = r.get('All State(s) Operating In', '').strip()
            hq_state = r.get('HQ State', '').strip()
            full_state = r.get('Full HQ State Name', '').strip()
            ownership = r.get('Ownership', '').strip()
            website = r.get('Website', '').strip()
            profit_type = r.get('Profit Type', '').strip()
            state_tier = r.get('State Tier', '').strip()
            contacts = r.get('Contacts', '').strip()
            override = r.get('Override', '').strip() == 'checked'

            companies[key] = {
                'name': name,
                'address': address,
                'allStates': all_states,
                'state': hq_state,
                'fullState': full_state,
                'ownership': ownership,
                'website': website,
                'profitType': profit_type,
                'stateTier': state_tier,
                'contacts': contacts,
                'override': override,
            }
    return companies


def resolve_state_abbr(state_str):
    """Given a state string (could be full name or abbreviation), return abbreviation."""
    s = state_str.strip()
    if len(s) == 2 and s.upper() in STATE_COORDS:
        return s.upper()
    if s in STATE_ABBREV:
        return STATE_ABBREV[s]
    return s


# ============================================================
# PARSE WORKING SHEET (company-level data)
# ============================================================
def parse_working_sheet(filepath):
    """Returns dict keyed by normalized company name."""
    companies = {}
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r.get('Name', '').strip()
            if not name:
                continue
            key = name.lower().strip()
            companies[key] = {
                'name': name,
                'address': r.get('HQ Address', '').strip(),
                'state': r.get('HQ State', '').strip(),
                'fullState': r.get('Full HQ State Name', '').strip(),
                'ownership': r.get('Ownership', '').strip(),
                'override': r.get('Override', '').strip() == 'checked',
                'website': r.get('Website', '').strip(),
                'allStates': r.get('All State(s) Operating In', '').strip(),
                'aiOwnership': r.get('AI Ownership Research', '').strip(),
                'contacts': r.get('Contacts', '').strip(),
                'profitType': r.get('Profit Type', '').strip(),
                'stateTier': r.get('State Tier', '').strip(),
            }
    return companies


# ============================================================
# PARSE GRID VIEW (message-level data) — THE KEY FILE
# ============================================================
def parse_grid_view(filepath):
    """
    Each row is a message. We aggregate per-company:
    - total messages sent, by medium
    - responded flag
    - scheduled intro call
    - assisted meeting
    - not interested
    - contacts reached
    - states, ownership, message types, accounts
    """
    company_data = defaultdict(lambda: {
        'msgsSent': 0,
        'byMedium': defaultdict(int),
        'byAccount': defaultdict(int),
        'responded': False,
        'respondedCount': 0,
        'scheduledIntro': False,
        'assistedMeeting': False,
        'notInterested': False,
        'followUpLater': False,
        'contacts': set(),
        'states': set(),
        'fullStates': set(),
        'ownership': '',
        'override': False,
        'stateTier': '',
        'meetingDate': '',
        'firstMsgDate': '',
        'lastMsgDate': '',
        'opened': False,
        'viewedProfile': False,
    })

    all_messages = []

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            company_raw = r.get('Companies (from Contacts)', '').strip()
            if not company_raw:
                continue

            # A message can be associated with multiple companies
            # Use the first one as primary
            company_names = [c.strip() for c in company_raw.split(',') if c.strip()]
            primary = company_names[0] if company_names else company_raw
            key = primary.lower().strip()

            cd = company_data[key]
            cd['name'] = primary
            cd['msgsSent'] += 1

            medium = r.get('Message Medium', '').strip()
            if medium:
                cd['byMedium'][medium] += 1

            account = r.get('Account', '').strip()
            if account:
                cd['byAccount'][account] += 1

            contact = r.get('Contacts', '').strip()
            if contact:
                cd['contacts'].add(contact)

            if r.get('Responded', '').strip() == 'checked':
                cd['responded'] = True
                cd['respondedCount'] += 1

            sched = r.get('Scheduled Intro Call', '').strip().upper()
            if sched == 'TRUE':
                cd['scheduledIntro'] = True

            assisted = r.get('Assisted Meeting', '').strip().upper()
            if assisted == 'TRUE':
                cd['assistedMeeting'] = True

            ni = r.get('Not Interested', '').strip().upper()
            if ni == 'NOT INTERESTED':
                cd['notInterested'] = True
            elif ni == 'FOLLOW UP LATER':
                cd['followUpLater'] = True

            # States
            states_raw = r.get('State(s) (from Companies) (from Contacts)', '').strip()
            if states_raw:
                for s in states_raw.split(','):
                    s = s.strip()
                    if s:
                        cd['states'].add(s)

            full_states = r.get('Full HQ State Name (from Companies) (from Contacts)', '').strip()
            if full_states:
                for s in full_states.split(','):
                    s = s.strip()
                    if s:
                        cd['fullStates'].add(s)

            own = r.get('Ownership (from Companies) (from Contacts)', '').strip()
            if own:
                # Take the first if comma-separated (multiple contacts)
                first_own = own.split(',')[0].strip()
                cd['ownership'] = first_own

            override_raw = r.get('Override (from Companies) (from Contacts)', '').strip()
            if 'checked' in override_raw.lower():
                cd['override'] = True

            tier = r.get('State Tier (from Companies) (from Contacts)', '').strip()
            if tier:
                cd['stateTier'] = tier.split(',')[0].strip()

            meeting_date = r.get('Meeting Date', '').strip()
            if meeting_date:
                cd['meetingDate'] = meeting_date

            if r.get('Opened', '').strip().upper() == 'TRUE':
                cd['opened'] = True
            if r.get('Viewed Profile', '').strip().upper() == 'TRUE':
                cd['viewedProfile'] = True

            date_sent = r.get('Date Sent', '').strip()
            if date_sent:
                if not cd['firstMsgDate'] or date_sent < cd['firstMsgDate']:
                    cd['firstMsgDate'] = date_sent
                if not cd['lastMsgDate'] or date_sent > cd['lastMsgDate']:
                    cd['lastMsgDate'] = date_sent

            all_messages.append({
                'company': primary,
                'contact': contact,
                'medium': medium,
                'account': account,
                'date': date_sent,
                'responded': r.get('Responded', '').strip() == 'checked',
            })

    return company_data, all_messages


# ============================================================
# PARSE PIPELINE EXCEL
# ============================================================
def parse_pipeline(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Pipeline Dashboard sheet
    ws = wb['Pipeline Dashboard']
    pipeline = []
    for row in range(7, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue

        states = str(ws.cell(row=row, column=3).value or '').strip()
        facility_type = str(ws.cell(row=row, column=4).value or '').strip()
        status = str(ws.cell(row=row, column=5).value or '').strip()
        priority = str(ws.cell(row=row, column=6).value or '').strip()
        ebitda = str(ws.cell(row=row, column=7).value or '').strip()
        asking_price = str(ws.cell(row=row, column=8).value or '').strip()
        nda_status = str(ws.cell(row=row, column=9).value or '').strip()
        data_room = str(ws.cell(row=row, column=10).value or '').strip()
        site_visit = str(ws.cell(row=row, column=11).value or '').strip()
        key_contact = str(ws.cell(row=row, column=12).value or '').strip()
        next_action = str(ws.cell(row=row, column=13).value or '').strip()
        action_owner = str(ws.cell(row=row, column=14).value or '').strip()
        deadline = str(ws.cell(row=row, column=15).value or '').strip()
        last_update = str(ws.cell(row=row, column=16).value or '').strip()
        days_since = ws.cell(row=row, column=17).value
        notes = str(ws.cell(row=row, column=18).value or '').strip()

        pipeline.append({
            'name': name,
            'states': states,
            'type': facility_type,
            'status': status,
            'priority': priority,
            'ebitda': ebitda,
            'askingPrice': asking_price,
            'ndaStatus': nda_status,
            'dataRoom': data_room,
            'siteVisit': site_visit,
            'keyContact': key_contact,
            'nextAction': next_action,
            'actionOwner': action_owner,
            'deadline': deadline,
            'lastUpdate': last_update,
            'daysSinceUpdate': int(days_since) if days_since else None,
            'notes': notes,
        })

    # Action tracker
    ws2 = wb['Action Tracker']
    actions = []
    for row in range(4, ws2.max_row + 1):
        priority = ws2.cell(row=row, column=1).value
        if not priority:
            continue
        actions.append({
            'priority': str(priority).strip(),
            'action': str(ws2.cell(row=row, column=2).value or '').strip(),
            'facility': str(ws2.cell(row=row, column=3).value or '').strip(),
            'owner': str(ws2.cell(row=row, column=4).value or '').strip(),
            'deadline': str(ws2.cell(row=row, column=5).value or '').strip(),
            'status': str(ws2.cell(row=row, column=6).value or '').strip(),
            'notes': str(ws2.cell(row=row, column=7).value or '').strip(),
            'pipelineStatus': str(ws2.cell(row=row, column=8).value or '').strip(),
        })

    return pipeline, actions


# ============================================================
# MERGE ALL DATA
# ============================================================
def merge_all():
    # 1. Parse View All for HQ addresses (most complete source)
    va_companies = parse_view_all('/Users/dylanpoler/Downloads/View All (3).csv')
    print(f"View All: {len(va_companies)} companies")

    # 2. Parse working sheet for supplemental info
    ws_companies = parse_working_sheet('/Users/dylanpoler/Downloads/Working Sheet (8).csv')
    print(f"Working Sheet: {len(ws_companies)} companies")

    # 3. Parse Grid view for complete message data
    gv_companies, all_messages = parse_grid_view('/Users/dylanpoler/Downloads/Grid view copy (1).csv')
    print(f"Grid View: {len(gv_companies)} companies, {len(all_messages)} messages")

    # 4. Parse pipeline
    pipeline, actions = parse_pipeline('/Users/dylanpoler/Downloads/Bradford_Pipeline_Dashboard (3).xlsx')
    print(f"Pipeline: {len(pipeline)} deals, {len(actions)} action items")

    # Build pipeline lookup by name
    pipeline_lookup = {}
    for p in pipeline:
        pipeline_lookup[p['name'].lower().strip()] = p

    # Alias map: pipeline deal names -> company keys (for fuzzy name mismatches)
    PIPELINE_ALIASES = {
        'nola detox & recovery center': 'nola detox',
        'the grove recovery': 'the grove recovery centers',
        'second chances': 'second chances addiction recovery center',
        'serenity treatment centers': 'serenity treatment center',
        'sanctuary': 'sanctuary louisiana',
        'turning leaf behavioral health': 'turning leaf behavioral health services',
        'asheville detox (healthcare alliance)': 'asheville detox center',
        'southeast detox / addiction ctr': 'southeast detox',
        'new waters': 'new waters recovery',
        'dreamlife / crestview': 'dreamlife recovery pa',
        'woodlake center': 'woodlake addiction recovery',
        'new vista / ethan crossing': 'ethan crossing addiction treatment',
        'cardinal': 'cardinal recovery',
        'the sylvia brafman mh center': 'the sylvia brafman mental health center',
        'peachtree detox (evoraa)': 'peachtree detox',
        'revive recover': 'gateway to sobriety (revive recover)',
        'southern sky': 'southern sky recovery',
        'the wave': 'the wave international',
        'recovery now / longbranch': 'longbranch healthcare',
        'sycamour': 'sycamore behavioral health',
        'ghr': 'ghr center for addiction recovery and treatment',
    }
    # Normalize unicode whitespace in pipeline_lookup keys
    normalized_lookup = {}
    for k, v in pipeline_lookup.items():
        norm_k = re.sub(r'\s+', ' ', k.replace('\u202f', ' ').replace('\u00a0', ' '))
        normalized_lookup[norm_k] = v
    pipeline_lookup.update(normalized_lookup)

    # Add aliased entries to pipeline_lookup so companies can find their pipeline deal
    for alias_from, alias_to in PIPELINE_ALIASES.items():
        if alias_from in pipeline_lookup and alias_to not in pipeline_lookup:
            pipeline_lookup[alias_to] = pipeline_lookup[alias_from]

    # ============================================================
    # GEOCODING — batch geocode all addresses via Census Bureau API
    # ============================================================
    geocode_cache = load_geocode_cache()
    print(f"Geocode cache: {len(geocode_cache)} entries")

    # Collect all unique addresses that need geocoding
    all_keys = set(gv_companies.keys()) | set(ws_companies.keys()) | set(va_companies.keys())
    addresses_to_geocode = []
    for key in all_keys:
        va = va_companies.get(key, {})
        ws = ws_companies.get(key, {})
        address = va.get('address', '') or ws.get('address', '')
        if address and key not in geocode_cache:
            addresses_to_geocode.append((key, address))

    if addresses_to_geocode:
        print(f"Geocoding {len(addresses_to_geocode)} new addresses via Census Bureau API...")
        geo_results = batch_geocode_census(addresses_to_geocode)
        for uid, coords in geo_results.items():
            geocode_cache[uid] = list(coords)
        # Mark failures so we don't retry
        for uid, _ in addresses_to_geocode:
            if uid not in geocode_cache:
                geocode_cache[uid] = None
        save_geocode_cache(geocode_cache)
        print(f"Geocoded {len(geo_results)} addresses successfully")
    else:
        print("All addresses already cached")

    # ============================================================
    # MERGE: Grid View is primary, enriched with View All + Working Sheet
    # ============================================================
    companies = []
    seen_keys = set()

    for key, gv in gv_companies.items():
        seen_keys.add(key)
        va = va_companies.get(key, {})
        ws = ws_companies.get(key, {})

        # Prefer View All address, then Working Sheet
        address = va.get('address', '') or ws.get('address', '')
        all_states_str = va.get('allStates', '') or ws.get('allStates', '') or ', '.join(sorted(gv['fullStates'] or gv['states']))

        state_abbr = va.get('state', '') or ws.get('state', '')
        if not state_abbr or state_abbr == '#ERROR!':
            for fs in gv['fullStates']:
                abbr = resolve_state_abbr(fs)
                if abbr in STATE_COORDS:
                    state_abbr = abbr
                    break
            if not state_abbr:
                for s in gv['states']:
                    abbr = resolve_state_abbr(s)
                    if abbr in STATE_COORDS:
                        state_abbr = abbr
                        break

        # Use geocoded coords if available, else fall back to state-level
        cached = geocode_cache.get(key)
        if cached:
            lat, lng = cached
        else:
            lat, lng = get_coords(address, state_abbr, gv['name'])
            if lat is None and gv['states']:
                for s in gv['states']:
                    abbr = resolve_state_abbr(s)
                    if abbr in STATE_COORDS:
                        lat, lng = get_coords('', abbr, gv['name'])
                        state_abbr = abbr
                        break

        norm_key = re.sub(r'\s+', ' ', key.replace('\u202f', ' ').replace('\u00a0', ' '))
        pipe = pipeline_lookup.get(key) or pipeline_lookup.get(norm_key)
        ownership = va.get('ownership', '') or ws.get('ownership', '') or gv['ownership']
        full_state = va.get('fullState', '') or ws.get('fullState', '') or (list(gv['fullStates'])[0] if gv['fullStates'] else '')

        companies.append({
            'name': gv['name'],
            'address': address,
            'state': state_abbr,
            'fullState': full_state,
            'ownership': ownership,
            'override': va.get('override', False) or ws.get('override', False) or gv['override'],
            'website': va.get('website', '') or ws.get('website', ''),
            'allStates': all_states_str,
            'stateTier': va.get('stateTier', '') or ws.get('stateTier', '') or gv['stateTier'],
            'msgsSent': gv['msgsSent'],
            'byMedium': dict(gv['byMedium']),
            'byAccount': dict(gv['byAccount']),
            'responded': gv['responded'],
            'respondedCount': gv['respondedCount'],
            'scheduledIntro': gv['scheduledIntro'],
            'assistedMeeting': gv['assistedMeeting'],
            'notInterested': gv['notInterested'],
            'followUpLater': gv['followUpLater'],
            'contactCount': len(gv['contacts']),
            'contacts': list(gv['contacts']),
            'firstMsg': gv['firstMsgDate'],
            'lastMsg': gv['lastMsgDate'],
            'meetingDate': gv['meetingDate'],
            'opened': gv['opened'],
            'viewedProfile': gv['viewedProfile'],
            'inPipeline': pipe is not None,
            'pipelineStatus': pipe['status'] if pipe else '',
            'pipelinePriority': pipe['priority'] if pipe else '',
            'pipelineType': pipe['type'] if pipe else '',
            'lat': lat,
            'lng': lng,
        })

    # Add companies from View All and Working Sheet not in Grid View
    for source in [va_companies, ws_companies]:
        for key, src in source.items():
            if key in seen_keys:
                continue
            seen_keys.add(key)
            va = va_companies.get(key, {})
            ws = ws_companies.get(key, {})

            address = va.get('address', '') or ws.get('address', '')
            state_abbr = va.get('state', '') or ws.get('state', '')
            if state_abbr == '#ERROR!':
                state_abbr = ''

            cached = geocode_cache.get(key)
            if cached:
                lat, lng = cached
            else:
                lat, lng = get_coords(address, state_abbr, src['name'])

            norm_key = re.sub(r'\s+', ' ', key.replace('\u202f', ' ').replace('\u00a0', ' '))
            pipe = pipeline_lookup.get(key) or pipeline_lookup.get(norm_key)

            companies.append({
                'name': src['name'],
                'address': address,
                'state': state_abbr,
                'fullState': va.get('fullState', '') or ws.get('fullState', ''),
                'ownership': va.get('ownership', '') or ws.get('ownership', ''),
                'override': va.get('override', False) or ws.get('override', False),
                'website': va.get('website', '') or ws.get('website', ''),
                'allStates': va.get('allStates', '') or ws.get('allStates', ''),
                'stateTier': va.get('stateTier', '') or ws.get('stateTier', ''),
                'msgsSent': 0,
                'byMedium': {},
                'byAccount': {},
                'responded': False,
                'respondedCount': 0,
                'scheduledIntro': False,
                'assistedMeeting': False,
                'notInterested': False,
                'followUpLater': False,
                'contactCount': 0,
                'contacts': [],
                'firstMsg': '',
                'lastMsg': '',
                'meetingDate': '',
                'opened': False,
                'viewedProfile': False,
                'inPipeline': pipe is not None,
                'pipelineStatus': pipe['status'] if pipe else '',
                'pipelinePriority': pipe['priority'] if pipe else '',
                'pipelineType': pipe['type'] if pipe else '',
                'lat': lat,
                'lng': lng,
            })

    # Compute message-level aggregates
    medium_counts = defaultdict(int)
    account_counts = defaultdict(int)
    monthly_counts = defaultdict(int)
    for m in all_messages:
        if m['medium']:
            medium_counts[m['medium']] += 1
        if m['account']:
            account_counts[m['account']] += 1
        if m['date']:
            # Extract month/year
            match = re.match(r'(\d{1,2})/\d{1,2}/(\d{4})', m['date'])
            if match:
                month_key = f"{match.group(2)}-{int(match.group(1)):02d}"
                monthly_counts[month_key] += 1

    meta = {
        'totalMessages': len(all_messages),
        'mediumCounts': dict(medium_counts),
        'accountCounts': dict(account_counts),
        'monthlyCounts': dict(sorted(monthly_counts.items())),
    }

    output = {
        'companies': companies,
        'pipeline': pipeline,
        'actions': actions,
        'meta': meta,
    }

    return output


if __name__ == '__main__':
    data = merge_all()
    with open('/Users/dylanpoler/Downloads/rehab_dashboard/data.json', 'w') as f:
        json.dump(data, f, indent=2)

    c = data['companies']
    gc = load_geocode_cache()
    geocoded = sum(1 for x in c if x['lat'] and gc.get(x['name'].lower().strip()))
    print(f"\n=== OUTPUT ===")
    print(f"Total companies: {len(c)}")
    print(f"With coordinates: {sum(1 for x in c if x['lat'])}")
    print(f"Geocoded (precise): {geocoded}")
    print(f"With allStates: {sum(1 for x in c if x['allStates'])}")
    print(f"With messages: {sum(1 for x in c if x['msgsSent'] > 0)}")
    print(f"Total messages: {data['meta']['totalMessages']}")
    print(f"Responded: {sum(1 for x in c if x['responded'])}")
    print(f"Scheduled intro: {sum(1 for x in c if x['scheduledIntro'])}")
    print(f"Assisted meeting: {sum(1 for x in c if x['assistedMeeting'])}")
    print(f"Not interested: {sum(1 for x in c if x['notInterested'])}")
    print(f"In pipeline: {sum(1 for x in c if x['inPipeline'])}")
    print(f"Pipeline deals: {len(data['pipeline'])}")
    print(f"Mediums: {data['meta']['mediumCounts']}")
    print(f"Accounts: {data['meta']['accountCounts']}")
